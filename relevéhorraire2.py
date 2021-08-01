import datetime
import openpyxl
from openpyxl import workbook



wb = openpyxl.load_workbook(r"C:\Users\konhotom\Desktop\Relevé_temps.xlsx")
pg1 = wb['Relevé_Temps']
pg1 =wb.active
t1 = datetime.datetime.strptime(input("Début du travail:"), '%H:%M')
t2 = datetime.datetime.strptime(input("Fin du travail:"), '%H:%M')
t3= t2-t1
print("Heure total:",t3)

row = pg1.max_row
for i in range(2,row+1):
    pg1.cell(row=i, column=1).value=t3

wb.save(r"C:\Users\konhotom\Desktop\Relevé_temps2.xlsx")
