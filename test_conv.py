import openpyxl
from openpyxl import Workbook
import docx
from docx import Document
from docx2pdf import convert
import os

doc = docx.Document(r'C:\Users\konhotom\Desktop\convoc.docx')
wb = openpyxl.load_workbook(r"C:\Users\konhotom\Desktop\convoc.xlsx")
pg1 = wb['pg1']
pg1 =wb.active
row = pg1.max_row
column = pg1.max_column

#Liste de prof récuperer sur excel
L_prof = []
for col_cells in pg1.iter_cols(min_col=0, max_col=column+1):
   for cell in col_cells:
       if(cell.value is not None):
           if 'M' in cell.value:
               L_prof.append(cell.value)
L = set(L_prof)
L_prof = list(L)
#print(L_prof)

'''
#test
count=0
while count < (pg1.max_column):
    for row in pg1.rows:
        if row[count].value == "M. Le Yeuch":
            print(row)
    count+=1
    print(row.value)

'''
#fonction permettant de cree les documents avec chaque nom de prof
def osef (L , z ,texts ):
    if z == 0:
        return 0
    else :
        nom = L_prof[z-1]
        for t in doc.paragraphs:
            inline = t.runs
            for x in range(len(inline)):
                if 'NOM' in inline[x].text:
                    text = inline[x].text.replace('NOM',nom )
                    inline[x].text = text
                    if text == L_prof[z-1]:
                        for z in range(z):
                            inline[x].text = L_prof[z-1]
                            print(t.text)
                            doc.save(f'Desktop\\test\\convoc_{z}.docx')
    return osef(L,z-1,texts)
osef(L_prof,len(L_prof),doc)

