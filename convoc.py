import openpyxl
from openpyxl import Workbook
import docx
from docx import Document
from docx2pdf import convert
import os
import six


doc = docx.Document(r'C:\Users\konhotom\Desktop\convoc.docx')
wb = openpyxl.load_workbook(r"C:\Users\konhotom\Desktop\convoc.xlsx")
pg1 = wb['pg1']
pg1 =wb.active
row = pg1.max_row
column = pg1.max_column
date = pg1.cell(row=1 , column=1).value




#Liste de prof récuperer sur excel
L_prof = []
for col_cells in pg1.iter_cols(min_col=0, max_col=column+1):
   for cell in col_cells:
       if(cell.value is not None):
           if 'M' in cell.value:
               L_prof.append(cell.value)
L = set(L_prof)
L_prof = list(L)
#print(L_prof)

def table_value():
    for i in range(1, row+1):
        for j in range(1,column+1):
            if(pg1.cell(i,j).value is not None):
                print(pg1.cell(i,j).value)


def table_index():
    for i in range(1, row+1):
        for j in range(1,column+1):
            if(pg1.cell(i,j).value is not None):
                print(pg1.cell(i,j).coordinate)


def find_value(elt):
    elt = elt
    for col_cells in pg1.iter_cols(min_col=0, max_col=column+1):
        for cell in col_cells:
            if(cell.value is not None):
                if elt in cell.value:
                    print(cell.value)
    return 0


def find_index(elt):
    elt = elt
    for col_cells in pg1.iter_cols(min_col=0, max_col=column+1):
        for cell in col_cells:
            if(cell.value is not None):
                if elt in cell.value:
                    print(cell.coordinate)
    return 0

def find_row(elt):
    elt = elt
    for row_cells in pg1.iter_rows(min_row=0, max_row=row+1):
        for cell in row_cells:
            if(cell.value is not None):
                if elt in cell.value:
                    print(cell.row)
    return 0
find_row('M. Gibaud')


def find_column(elt):
    elt = elt
    for col_cells in pg1.iter_cols(min_col=0, max_col=column+1):
        for cell in col_cells:
            if(cell.value is not None):
                if elt in cell.value:
                    print(cell.column)
    return 0


def index_prof (L , z  ):
    if z == 0:
        return 0
    else :
        nom = L_prof[z-1]
        find_value(nom)
    return index_prof(L,z-1)



def first_cell_row(elt):
    elt = elt
    ligne = pg1['A']
    for col_cells in pg1.iter_cols(min_col=0, max_col=column+1):
        for cell in col_cells:
            if(cell.value is not None):
                if elt in cell.value:
                    print()

    return 0
first_cell_row('M. Gibaud')








heure = pg1['A']

for heure1 in range(1,len(heure)):
    if(heure[heure1].value is not None):
        if '0' in heure[heure1].value:
            print(heure[heure1].value)

salle = pg1[2]

for salle1 in range(1,len(salle)):
    if(salle[salle1].value is not None):
        print(salle  [salle1].value)


#fonction permettant de cree les documents avec chaque nom de prof
def remplacer (L , z ,texts ):
    if z == 0:
        return 0
    else :
        nom = L_prof[z-1]
        for t in doc.paragraphs:
            inline = t.runs
            for x in range(len(inline)):
                if 'NOM' in inline[x].text:
                    text = inline[x].text.replace('NOM',nom )
                    inline[x].text = text
                    if text == L_prof[z-1]:
                        for z in range(z):
                            inline[x].text = L_prof[z-1]
                            #doc.save(f'Desktop\\test\\convoc_{z}.docx')
    return remplacer(L,z-1,texts)
remplacer(L_prof,len(L_prof),doc)


def creer(L,z):
    if z ==0 :
        return 0
    else:
        remplacer(L_prof,len(L_prof),doc)
        doc.save(f'Desktop\\test\\convoc_{z}.docx')
    return creer(L,z-1)
#creer(L,len(L_prof))


