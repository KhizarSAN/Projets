# Commencer par tapper dans la console 'pip install openpyxl' puis importer le
import openpyxl

# Créer une classe quelconque suivi de 'openpyxl.load_worknook' et placer le path du fichier dans l'argument (attention il faut mettre deux \)
#C'est a mettre au début de CHAQUE CODE concernant cette page
wb = openpyxl.load_workbook("C:\\Users\\konhotom\\Desktop\\testexcel.xlsx")
pages = wb.sheetnames


#Pour savoir sur quel page on se situe on tappe 'print(wb.active.title)'
#print(wb.active.title)
#Pour afficher le contenue d'une case il y'a plusieurs méthode :


#Méthode 1(basique)
pg1 = wb['test1']
data = pg1['A1'].value
print(data)

#Méthode 2 (rapide)
print(pg1.cell(3,3).value)

#Méthode 3 (précis pour dire sur quel page on travaille)
print(wb['test1']['A3'].value)

#Pour changer de page ou lire le contenue d'une autre page on définit d'abord celle-ci :
#pg2 = wb['nom de la page ']

#Puis on tappe la meme commande que pour la pg1 mais en la remplacant par la 2
# print(pg2.cell(1,3).value)

# Pour afficher le contenu de TOUTES les cellules on nommera d'abord la page  :
sh1 = wb['test1']
# Ensuite on va définir ce qu'on veut en l'occurence le nombre max de colonnes et de ligne :

row = pg1.max_row
column = pg1.max_column

# Et enfin faire un print d'eux tous ( attention mettre un +1 dans l'argument sinon il manquera un contenu de chaque)
for i in range(1, row+1):
    for j in range(1,column+1):
        print(pg1.cell(i,j).value)

# Pour écrire dans le fichier il suffit de mettre les coordonnées de la cellule en question et les remplir :

pg1.cell(row = 1, column=4 , value = 'XVBARBAR')
pg1.cell(row = 2, column=4 , value = '20 ans')
pg1.cell(row = 3, column=4 , value =  'rappeurs')

# Puis faire une sauvegarde (ca marche mieux sur pycharm)
wb.save("testexcel2.xlsx")