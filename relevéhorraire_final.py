import datetime
import openpyxl
from openpyxl import workbook



wb = openpyxl.load_workbook(r"C:\Users\konhotom\Desktop\relevé_test.xlsx")
pg1 = wb['Relevé_Temps']
pg1 =wb.active
row = pg1.max_row

#Input date
year = int(input('Enter a year'))
day = int(input('Enter a day'))
month = int(input('Enter a month'))
date1 = datetime.date(year, month, day)
print(date1)

if pg1.cell(row=row , column=1).value == None :
    pg1.cell(row=row, column=1).value=date1
else:
        row + 1
        pg1.cell(row=row +1, column=1).value=date1
#Input horraire
t1 = datetime.datetime.strptime(input("Début du travail:"), '%H:%M')
t2 = datetime.datetime.strptime(input("Fin du travail:"), '%H:%M')
t3= t2-t1
tps_un = t1
tps_deux= t2
print("Heure total:",t3)

temps_1_case = str(t1.hour) + ":" + str(t1.minute)
temps_2_case = str(t2.hour) + ":" + str(t2.minute)

#Total travail
if pg1.cell(row=row , column=5).value == None :
    pg1.cell(row=row, column=5).value=t3
else:
        row + 1
        pg1.cell(row=row +1, column=5).value=t3

#Début travail
if pg1.cell(row=row , column=2).value == None :
    pg1.cell(row=row, column=2).value=temps_1_case
else:
        row + 1
        pg1.cell(row=row +1, column=2).value=temps_1_case

#Fin travail
if pg1.cell(row=row , column=3).value == None :
    pg1.cell(row=row, column=3).value=temps_2_case
else:
        row + 1
        pg1.cell(row=row +1, column=3).value=temps_2_case

wb.save(r"C:\Users\konhotom\Desktop\relevé_test.xlsx")
